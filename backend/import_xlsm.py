"""Міграція даних з реальної книги ЗБЛ (*.xlsm) у базу даних додатка.

Читає таблиці Excel через openpyxl і завантажує:
  * Довідники!TabAbon  -> houses (реєстр будинків/абонентів)
  * Довідники!TabTN    -> heat_load (теплове навантаження, join за адресою)
  * Архів!TabBaza      -> readings (історія показань)

Використання:
    python import_xlsm.py "/шлях/до/ЗБЛ новий сезон.xlsm"

Скрипт стійкий до відсутніх значень: порожні/некоректні клітинки пропускаються.
Це стартова версія Етапу 0 — за потреби маппінг колонок легко уточнити.
"""
from __future__ import annotations

import sys
from datetime import date, datetime

from openpyxl import load_workbook

from app.database import Base, SessionLocal, engine
from app import models


def _num(v, default=0.0) -> float:
    try:
        if v in (None, "", "-"):
            return default
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (TypeError, ValueError):
        return default


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _table_rows(ws, table_name):
    """Повертає (заголовки, рядки) іменованої таблиці аркуша."""
    tbl = ws.tables[table_name]
    cells = ws[tbl.ref]
    rows = [[c.value for c in row] for row in cells]
    return rows[0], rows[1:]


def import_workbook(path: str):
    Base.metadata.create_all(bind=engine)
    wb = load_workbook(path, data_only=True, read_only=True)
    db = SessionLocal()
    try:
        ws_ref = wb["Довідники"]
        ws_arh = wb["Архів"]

        # --- Теплове навантаження: адреса -> Гкал/год ---
        load_by_addr: dict[str, float] = {}
        try:
            _, tn_rows = _table_rows(ws_ref, "TabTN")
            for r in tn_rows:
                if r and r[0]:
                    load_by_addr[str(r[0]).strip()] = _num(r[1])
        except KeyError:
            print("Попередження: таблицю TabTN не знайдено, навантаження = 0.")

        # --- Котельні (унікальні з TabAbon, col16) ---
        boilers: dict[str, models.Boiler] = {}

        def boiler_id(name):
            name = (name or "").strip()
            if not name:
                return None
            if name not in boilers:
                b = models.Boiler(name=name)
                db.add(b)
                db.flush()
                boilers[name] = b
            return boilers[name].id

        # --- Будинки (TabAbon) ---
        houses: dict[str, models.House] = {}
        _, ab_rows = _table_rows(ws_ref, "TabAbon")
        for r in ab_rows:
            code = (str(r[0]).strip() if r and r[0] else "")
            if not code:
                continue
            address = (str(r[3]).strip() if len(r) > 3 and r[3] else "")
            house = models.House(
                code=code,
                personal_account=str(r[1]).strip() if len(r) > 1 and r[1] else None,
                owner=str(r[2]).strip() if len(r) > 2 and r[2] else None,
                address=address or None,
                meter_no=str(r[4]).strip() if len(r) > 4 and r[4] else None,
                meter_type=str(r[5]).strip() if len(r) > 5 and r[5] else None,
                unit=(str(r[7]).strip() if len(r) > 7 and r[7] else "Гкал"),
                settlement=str(r[14]).strip() if len(r) > 14 and r[14] else None,
                boiler_id=boiler_id(r[15] if len(r) > 15 else None),
                commercial_metering=bool(r[16]) if len(r) > 16 else True,
                heat_load=load_by_addr.get(address, 0.0),
                share=_num(r[25], 1.0) if len(r) > 25 else 1.0,
            )
            db.add(house)
            houses[code] = house
        db.flush()
        print(f"Імпортовано будинків: {len(houses)}")

        # --- Показання (TabBaza) ---
        _, bz_rows = _table_rows(ws_arh, "TabBaza")
        count = 0
        for r in bz_rows:
            code = (str(r[2]).strip() if len(r) > 2 and r[2] else "")
            house = houses.get(code)
            if house is None:
                continue
            close_field = r[23] if len(r) > 23 else None
            is_control = isinstance(close_field, str) and "Контрольне" in close_field
            energy_current = _num(r[7]) if len(r) > 7 else 0.0
            delta = _num(r[8]) if len(r) > 8 else 0.0
            db.add(
                models.Reading(
                    house_id=house.id,
                    measure_date=None if is_control else _to_date(r[4] if len(r) > 4 else None),
                    is_control=is_control,
                    duration_hours=_num(r[6]) if len(r) > 6 else 0.0,
                    t_avg=_num(r[24]) if len(r) > 24 else 0.0,
                    unit=str(r[9]).strip() if len(r) > 9 and r[9] else house.unit,
                    energy_current=energy_current,
                    energy_previous=round(energy_current - delta, 3),
                    energy_delta_units=delta,
                    volume_gcal=_num(r[10]) if len(r) > 10 else 0.0,
                    q_calculated=_num(r[25]) if len(r) > 25 else 0.0,
                    difference=_num(r[26]) if len(r) > 26 else 0.0,
                    percent=_num(r[27]) if len(r) > 27 else None,
                )
            )
            count += 1
        db.commit()
        print(f"Імпортовано показань: {count}")
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Вкажіть шлях до .xlsm: python import_xlsm.py файл.xlsm")
        sys.exit(1)
    import_workbook(sys.argv[1])
